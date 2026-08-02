import pandas as pd

CAMINHO_MATRICULA = r"\Tabela_Matricula_2025.csv"
CAMINHO_ESCOLA = r"\Tabela_Escola_2025.csv"
ARQUIVO_SAIDA = r"\matriculas_creche_rmrj_2025.xlsx"

SEP = ";"
ENCODING = "latin-1"

MAPA_DEPENDENCIA = {
    1: "Federal",
    2: "Estadual",
    3: "Municipal",
    4: "Privada",
}

MUNICIPIOS_RMRJ = [
    "Belford Roxo",
    "Cachoeiras de Macacu",
    "Duque de Caxias",
    "Guapimirim",
    "Itaboraí",
    "Itaguaí",
    "Japeri",
    "Magé",
    "Maricá",
    "Mesquita",
    "Nilópolis",
    "Niterói",
    "Nova Iguaçu",
    "Paracambi",
    "Petrópolis",
    "Queimados",
    "Rio Bonito",
    "Rio de Janeiro",
    "São Gonçalo",
    "São João de Meriti",
    "Seropédica",
    "Tanguá"
]

UF_RJ = "RJ"

def carregar_dados():
    matricula = pd.read_csv(
        CAMINHO_MATRICULA,
        sep=SEP,
        encoding=ENCODING,
        usecols=["CO_ENTIDADE", "QT_MAT_INF_CRE"],
    )
    escola = pd.read_csv(
        CAMINHO_ESCOLA,
        sep=SEP,
        encoding=ENCODING,
        usecols=["CO_ENTIDADE", "SG_UF", "NO_MUNICIPIO", "TP_DEPENDENCIA"],
    )

    df = matricula.merge(escola, on="CO_ENTIDADE", how="inner")

    # Garante que a coluna de matrículas seja numérica (nulos viram 0)
    df["QT_MAT_INF_CRE"] = pd.to_numeric(
        df["QT_MAT_INF_CRE"], errors="coerce"
    ).fillna(0)

    # Nome legível da dependência administrativa
    df["DEPENDENCIA"] = df["TP_DEPENDENCIA"].map(MAPA_DEPENDENCIA)

    return df


def agregar_territorio(df, nome_territorio):
    """Recebe um recorte do DataFrame e devolve uma linha (dict) com o
    total e o detalhamento por dependência administrativa."""

    linha = {"Território": nome_territorio}
    linha["Total"] = df["QT_MAT_INF_CRE"].sum()

    for dep in MAPA_DEPENDENCIA.values():
        linha[dep] = df.loc[df["DEPENDENCIA"] == dep, "QT_MAT_INF_CRE"].sum()

    return linha


def montar_planilha(df):
    linhas = []

    df_rmrj = df[
        (df["SG_UF"] == UF_RJ) & (df["NO_MUNICIPIO"].isin(MUNICIPIOS_RMRJ))
    ]

    for municipio in MUNICIPIOS_RMRJ:
        df_mun = df_rmrj[df_rmrj["NO_MUNICIPIO"] == municipio]
        linhas.append(agregar_territorio(df_mun, municipio))

    linhas.append(agregar_territorio(df_rmrj, "RMRJ (Total)"))

    df_uf = df[df["SG_UF"] == UF_RJ]
    linhas.append(agregar_territorio(df_uf, "Estado do Rio de Janeiro"))

    linhas.append(agregar_territorio(df, "Brasil"))

    colunas = ["Território", "Total"] + list(MAPA_DEPENDENCIA.values())
    resultado = pd.DataFrame(linhas, columns=colunas)

    for col in colunas[1:]:
        resultado[col] = resultado[col].astype(int)

    return resultado


def exportar_xlsx(resultado, caminho_saida):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Matrículas Creche RMRJ"

    fonte_padrao = "Arial"

    # Cabeçalho
    for j, coluna in enumerate(resultado.columns, start=1):
        cel = ws.cell(row=1, column=j, value=coluna)
        cel.font = Font(name=fonte_padrao, bold=True, color="FFFFFF")
        cel.fill = PatternFill("solid", fgColor="305496")
        cel.alignment = Alignment(horizontal="center")

    # Linhas de dados
    linhas_destaque = {"RMRJ (Total)", "Estado do Rio de Janeiro", "Brasil"}
    for i, row in enumerate(resultado.itertuples(index=False), start=2):
        for j, valor in enumerate(row, start=1):
            cel = ws.cell(row=i, column=j, value=valor)
            cel.font = Font(
                name=fonte_padrao,
                bold=(row.Território in linhas_destaque),
            )
            if j > 1:
                cel.number_format = "#,##0"
                cel.alignment = Alignment(horizontal="right")

    # Largura das colunas
    larguras = [26, 12, 12, 12, 12, 12]
    for j, largura in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(j)].width = largura

    ws.freeze_panes = "A2"

    wb.save(caminho_saida)


if __name__ == "__main__":
    df = carregar_dados()
    resultado = montar_planilha(df)
    exportar_xlsx(resultado, ARQUIVO_SAIDA)
    print(f"Planilha gerada: {ARQUIVO_SAIDA}")
    print(resultado)